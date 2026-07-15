import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Any
import logging
import sys

# Configuración básica de logging para informar el progreso o errores
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

def escribir_celda(ws: Worksheet, coord: str, valor: Any) -> None:
    """
    Escribe un valor en una celda de manera segura. Si la celda es parte de un rango
    combinado (MergedCell), encuentra automáticamente la celda principal (superior izquierda)
    de ese rango y escribe el valor allí para evitar el error 'MergedCell is read-only'.
    """
    celda = ws[coord]
    if type(celda).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                top_left_coord = str(merged_range).split(':')[0]
                ws[top_left_coord] = valor
                return
    else:
        ws[coord] = valor

def llenar_hoja_datos(ws: Worksheet, datos_cliente: Dict[str, str]) -> None:
    """
    Inyecta los datos generales del cliente en la hoja 'DATOS '.
    """
    logging.info("Llenando hoja 'DATOS '...")
    # TODO: Actualizar las coordenadas ('C4', 'C5', etc.) con las celdas reales de tu plantilla "DATOS "
    escribir_celda(ws, 'C4', datos_cliente.get('razon_social', ''))
    escribir_celda(ws, 'C5', datos_cliente.get('cuit', ''))
    escribir_celda(ws, 'C6', datos_cliente.get('direccion', ''))
    escribir_celda(ws, 'C7', datos_cliente.get('localidad', ''))
    escribir_celda(ws, 'C8', datos_cliente.get('provincia', ''))
    escribir_celda(ws, 'C9', datos_cliente.get('cp', ''))

def llenar_hoja_grilla(ws: Worksheet, mediciones_pat: List[Dict[str, Any]]) -> None:
    """
    Inyecta las mediciones de Puesta a Tierra en la hoja 'GRILLA' (y similares).
    """
    logging.info(f"Llenando hoja '{ws.title}'...")
    # TODO: Actualizar la fila inicial (ej. fila_inicio = 12) según tu plantilla
    fila_inicio = 12
    
    for i, medicion in enumerate(mediciones_pat):
        fila_actual = fila_inicio + i
        # TODO: Actualizar las letras de las columnas ('A', 'B', 'C', etc.) según tu plantilla
        escribir_celda(ws, f'A{fila_actual}', medicion.get('nro_toma', ''))
        escribir_celda(ws, f'B{fila_actual}', medicion.get('sector', ''))
        escribir_celda(ws, f'C{fila_actual}', medicion.get('condicion_terreno', ''))
        escribir_celda(ws, f'D{fila_actual}', medicion.get('valor_ohms', ''))
        escribir_celda(ws, f'E{fila_actual}', medicion.get('cumple', ''))

def llenar_hojas_diferenciales(ws: Worksheet, mediciones_dif: List[Dict[str, Any]]) -> None:
    """
    Inyecta las mediciones de Disyuntores Diferenciales en hojas como 'DIFERENCIALES'.
    """
    logging.info(f"Llenando hoja '{ws.title}'...")
    # TODO: Actualizar la fila inicial (ej. fila_inicio = 15) según tu plantilla
    fila_inicio = 15
    
    for i, medicion in enumerate(mediciones_dif):
        fila_actual = fila_inicio + i
        # TODO: Actualizar las letras de las columnas ('A', 'B', 'C', etc.) según tu plantilla
        escribir_celda(ws, f'A{fila_actual}', medicion.get('sector', ''))
        escribir_celda(ws, f'B{fila_actual}', medicion.get('marca', ''))
        escribir_celda(ws, f'C{fila_actual}', medicion.get('corriente_nominal_a', ''))
        escribir_celda(ws, f'D{fila_actual}', medicion.get('corriente_fuga_ma', ''))
        escribir_celda(ws, f'E{fila_actual}', medicion.get('tiempo_disparo_ms', ''))
        escribir_celda(ws, f'F{fila_actual}', medicion.get('cumple', ''))

def llenar_hoja_conclusiones(ws: Worksheet, conclusiones: Dict[str, str]) -> None:
    """
    Inyecta las conclusiones, observaciones y datos del equipo en la hoja 'CONCLUSIONES'.
    """
    logging.info("Llenando hoja 'CONCLUSIONES'...")
    # TODO: Actualizar las coordenadas con las celdas reales de tu plantilla "CONCLUSIONES"
    escribir_celda(ws, 'B10', conclusiones.get('observaciones', ''))
    escribir_celda(ws, 'B15', conclusiones.get('instrumento_utilizado', ''))
    escribir_celda(ws, 'F4', conclusiones.get('fecha_medicion', ''))

def generar_protocolo(ruta_plantilla: str, ruta_salida: str, datos: Dict[str, Any]) -> None:
    """
    Función principal que orquesta la lectura, llenado y guardado del protocolo Excel.
    """
    try:
        logging.info(f"Intentando abrir la plantilla: {ruta_plantilla}")
        # Cargar el libro manteniendo los formatos (data_only=False preserva las fórmulas, pero para 
        # mantener estilos generales no es estrictamente necesario pasarlo, por defecto lee todo)
        wb: Workbook = openpyxl.load_workbook(ruta_plantilla)
        
        # 1. Hoja DATOS
        if 'DATOS ' in wb.sheetnames:
            llenar_hoja_datos(wb['DATOS '], datos['datos_cliente'])
        else:
            logging.warning("No se encontró la hoja 'DATOS ' en la plantilla.")

        # 2. Hojas de GRILLA (Puesta a Tierra)
        hojas_grilla = ['GRILLA', 'GRILLA (3)', 'GRILLA (4)']
        for nombre_hoja in hojas_grilla:
            if nombre_hoja in wb.sheetnames:
                # Aquí podrías dividir la lista de mediciones_pat si son demasiadas para una sola hoja
                # Por ahora le pasamos la lista completa a todas como ejemplo.
                llenar_hoja_grilla(wb[nombre_hoja], datos['mediciones_pat'])
            else:
                logging.warning(f"No se encontró la hoja '{nombre_hoja}' en la plantilla.")

        # 3. Hojas de DIFERENCIALES
        hojas_diferenciales = ['DIF. TRI', 'DIFER. COMB.', 'DIFERENCIALES']
        for nombre_hoja in hojas_diferenciales:
            if nombre_hoja in wb.sheetnames:
                llenar_hojas_diferenciales(wb[nombre_hoja], datos['mediciones_diferenciales'])
            else:
                logging.warning(f"No se encontró la hoja '{nombre_hoja}' en la plantilla.")

        # 4. Hoja CONCLUSIONES
        if 'CONCLUSIONES' in wb.sheetnames:
            llenar_hoja_conclusiones(wb['CONCLUSIONES'], datos['conclusiones'])
        else:
            logging.warning("No se encontró la hoja 'CONCLUSIONES' en la plantilla.")

        # Guardar como un archivo nuevo
        logging.info(f"Guardando el protocolo generado en: {ruta_salida}")
        wb.save(ruta_salida)
        logging.info("¡Proceso completado con éxito!")

    except FileNotFoundError:
        logging.error(f"Error: No se encontró el archivo de plantilla en la ruta: {ruta_plantilla}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Ocurrió un error inesperado al procesar el Excel: {e}")
        sys.exit(1)

