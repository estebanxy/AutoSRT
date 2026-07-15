import openpyxl
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:D4')
ws['B2'] = 'Initial'

def set_val(ws, coord, val):
    cell = ws[coord]
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                top_left = str(merged_range).split(':')[0]
                ws[top_left] = val
                return
    else:
        ws[coord] = val

set_val(ws, 'C3', 'New Value')
print(ws['B2'].value)
